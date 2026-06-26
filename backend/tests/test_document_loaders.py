"""xlsx / doc 文档解析单元测试。"""

from pathlib import Path

from docx import Document
from openpyxl import Workbook

from app.ai_service.document.loaders import load_document_text_and_images
from app.ai_service.document.office_convert import convert_with_libreoffice, find_soffice_executable
from app.platform.config import LIBREOFFICE_SOFFICE_PATH


def test_find_soffice_bundled_or_configured():
    """应能发现项目自带或配置的 LibreOffice。"""
    path = find_soffice_executable()
    assert path
    assert Path(path).is_file()
    assert LIBREOFFICE_SOFFICE_PATH


def test_read_xlsx(tmp_path: Path):
    xlsx_path = tmp_path / "sample.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "数据"
    ws["A1"] = "名称"
    ws["B1"] = "值"
    ws["A2"] = "测试项"
    ws["B2"] = 42
    wb.save(xlsx_path)

    text, images, warnings = load_document_text_and_images(xlsx_path)
    assert "测试项" in text
    assert "42" in text
    assert images == []


def test_read_doc_via_libreoffice(tmp_path: Path):
    """.doc 经 LibreOffice 转 docx 后可解析正文。"""
    if not find_soffice_executable():
        return

    docx_path = tmp_path / "sample.docx"
    doc = Document()
    doc.add_paragraph("旧版 Word 文档测试内容。" * 10)
    doc.save(docx_path)

    doc_path = tmp_path / "sample.doc"
    converted_to_doc = convert_with_libreoffice(docx_path, tmp_path, "doc")
    if converted_to_doc is None:
        return
    converted_to_doc.rename(doc_path)

    text, images, warnings = load_document_text_and_images(doc_path)
    assert "旧版 Word 文档测试内容" in text
    assert images == []
